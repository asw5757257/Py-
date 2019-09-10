from openpyxl import Workbook
from openpyxl import load_workbook
import os
import pymysql

"""
ItemUtils 能够以表的形式存储内容，并保存到 Excel 文档和数据库中
郭钰鸣 20168184
2019.09.01
"""


class Item:
    def __init__(self, name, content, score, say):
        self.name = name
        self.content = content
        self.score = score
        self.say = say

    def __str__(self):
        string = "name=" + self.name + \
                 ", content=" + self.content + \
                 ", score=" + self.score + \
                 ", say=" + self.say
        return string


class ItemUtils:
    list = []

    def append_item(self, item):
        self.list.append(item)

    def append(self, name, content, score, say):
        insert_item = Item(name, content, score, say)
        self.append_item(insert_item)

    def clear(self):
        self.list.clear()

    def save_excel(self, sheet="工作表", file="list.xlsx"):
        if os.path.exists(file):
            print("文件存在，打开")
        else:
            print("文件不存在，新建")
            wb = Workbook()
            wb.save(file)

        wb = load_workbook(file)
        sheets = wb.sheetnames
        if sheet in sheets:
            wb.remove(wb[sheet])

        wb.create_sheet(sheet)
        wb.active = wb[sheet]
        ws = wb.active
        ws.append(["名称","内容","评分","一句话评语"])
        for item in self.list:
            ws.append([item.name, item.content, item.score, item.say])

        wb.save(file)
        print("文件已保存。")

    def save_database(self, first=False):
        """保存到数据库，如果 first==True 就新建表"""
        sql_drop = """drop table if exists `douban`"""
        sql_create = """
        CREATE TABLE  `douban` (
          `ID` int(10) NOT NULL AUTO_INCREMENT,
          `NAME` varchar(250) COLLATE utf8_bin NOT NULL,
          `CONTENT` varchar(250) COLLATE utf8_bin NOT NULL,
          `SCORE` FLOAT(5) COLLATE utf8_bin NOT NULL,
          `SAY` varchar(250) COLLATE utf8_bin NOT NULL,
          PRIMARY KEY (`ID`)
        )ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_bin
        """
        sql_insert_format = "insert into `dbdx`.`douban` (`NAME`, `CONTENT`, `SCORE`, `SAY`) values ('%s', '%s', '%s', '%s')"

        conn = pymysql.connect(host="localhost", user="root", passwd="root", port=3306, db="dbdx", charset="utf8")

        try:
            with conn.cursor() as cursor:
                cursor.execute(sql_drop)
            conn.commit()

            with conn.cursor() as cursor:
                cursor.execute(sql_create)
            conn.commit()

            with conn.cursor() as cursor:
                for item in self.list:
                    print(sql_insert_format % (item.name, item.content, item.score, item.say))
                    cursor.execute(sql_insert_format % (item.name, item.content, item.score, item.say))
                conn.commit()
        finally:
            conn.close()
        print("数据库保存结束")

    def __str__(self):
        string = "list=["
        for i in range(0, len(self.list)):
            string += "[" + str(self.list[i]) + "]"
            if i != len(self.list) - 1:
                string += ",\n"
        string += "]"
        return string


"""
if __name__ == '__main__':
    # 实例化一个工具类
    item_list = ItemUtils()

    # 添加一个项目到列表中
    # 注意：字符串中不能含有换行符！
    item_list.append("肖申克的救赎", "导演: 弗兰克·德拉邦特", "9.1", "希望让人自由。")

    # 输出列表内的所有内容
    print(item_list)

    # 将列表中的项目保存到 Excel 文件
    # 可以指定表名和保存文件的名字，没有参数时使用默认值 (sheet="工作表", file="list.xlsx")
    # 相同表名的内容会被覆盖
    item_list.save_excel(sheet="表", file="list.xlsx")
    
    # 将列表中的所有项目插入到数据库中
    # first=True 表示是第一次执行，会运行建表命令，没有参数时 first=False
    item_list.save_database(first=True)

    # 清空列表
    item_list.clear()
    print(item_list)
"""
